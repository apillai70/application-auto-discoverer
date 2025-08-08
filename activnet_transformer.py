#!/usr/bin/env python3
"""
ACTIVnet Data Transformer
Extracts data from ACTIVnet_Sample.xlsx and transforms it to match the format of 
synthetic_flows_apps_archetype_mapped.xlsx with comprehensive port/service mapping.
"""

import pandas as pd
import re
from datetime import datetime
import openpyxl
from openpyxl import load_workbook

def create_port_service_mapping():
    """
    Create a comprehensive mapping of ports and protocols to their services.
    Based on IANA registry, common port assignments, and research.
    """
    port_services = {
        # Well-known ports (0-1023)
        7: "Echo Protocol",
        21: "FTP (File Transfer Protocol)",
        22: "SSH (Secure Shell)",
        23: "Telnet",
        25: "SMTP (Simple Mail Transfer Protocol)",
        37: "Time Protocol",
        53: "DNS (Domain Name System)",
        67: "DHCP Server (Bootstrap Protocol)",
        68: "DHCP Client (Bootstrap Protocol)",
        80: "HTTP (HyperText Transfer Protocol)",
        88: "Kerberos",
        110: "POP3 (Post Office Protocol v3)",
        119: "NNTP (Network News Transfer Protocol)",
        123: "NTP (Network Time Protocol)",
        135: "Microsoft RPC (Remote Procedure Call)",
        137: "NetBIOS Name Service",
        138: "NetBIOS Datagram Service",
        139: "NetBIOS Session Service",
        143: "IMAP (Internet Message Access Protocol)",
        161: "SNMP (Simple Network Management Protocol)",
        177: "XDMCP (X Display Manager Control Protocol)",
        389: "LDAP (Lightweight Directory Access Protocol)",
        443: "HTTPS (HTTP over SSL/TLS)",
        445: "SMB (Server Message Block)",
        465: "SMTPS (SMTP over SSL)",
        514: "Syslog",
        515: "LPR/LPD (Line Printer Daemon)",
        523: "IBM DB2 Database Server",
        631: "IPP (Internet Printing Protocol)",
        993: "IMAPS (IMAP over SSL)",
        995: "POP3S (POP3 over SSL)",
        
        # Registered ports (1024-49151)
        1433: "Microsoft SQL Server",
        1521: "Oracle Database",
        1723: "PPTP (Point-to-Point Tunneling Protocol)",
        3231: "VidiGo Communication",
        3306: "MySQL Database",
        3389: "RDP (Remote Desktop Protocol)",
        5060: "SIP (Session Initiation Protocol)",
        5432: "PostgreSQL Database",
        8080: "HTTP Alternate/Proxy",
        8090: "HTTP Alternate",
        8443: "HTTPS Alternate",
        9100: "HP JetDirect / PDL Data Stream / Raw Network Printing",
        
        # Higher numbered ports (commonly used)
        10247: "Application Server",
        10355: "Application Server",
        10518: "Application Server",
        10648: "Application Server",
        11827: "Application Server",
        12106: "Application Server",
        12270: "Application Server",
        12427: "Application Server",
        13335: "Application Server",
        16107: "Application Server",
        16119: "Application Server",
        16121: "Application Server",
        16131: "Application Server",
        16138: "Application Server",
        16144: "Application Server",
        16147: "Application Server",
        16156: "Application Server",
        16167: "Application Server",
        16170: "Application Server",
        16171: "Application Server",
        16184: "Application Server",
        
        # Gaming and specialized ports
        26001: "Gaming/Application Server",
        26198: "Unassigned (Private Use)",
        27019: "Steam Gaming",
        27025: "Half-Life Gaming",
        27961: "Quake III Gaming",
        30721: "Application Server",
        35692: "Application Server",
        49667: "Dynamic/Private Port",
        
        # Other specialized ports
        5297: "Application Server",
        5474: "Application Server",
        5496: "Application Server",
        5773: "Application Server",
        5795: "Application Server",
        6412: "Application Server",
        6420: "Application Server",
        6705: "Application Server",
        6783: "Application Server",
        6974: "Application Server",
        7054: "Application Server",
        7076: "Application Server",
        7311: "Application Server",
        7990: "Application Server",
        8122: "Application Server",
        8329: "Application Server",
        8725: "Application Server",
        9206: "Application Server",
        9439: "Application Server",
        9467: "Application Server",
    }
    
    # Protocol mappings for non-port protocols
    protocol_services = {
        "DNS": "Domain Name System",
        "HTTP": "HyperText Transfer Protocol",
        "HTTP2-SSL": "HTTP/2 over SSL/TLS",
        "HTTPS": "HTTP over SSL/TLS",
        "HTTps": "HTTP over SSL/TLS (variant)",
        "ICMPv6": "Internet Control Message Protocol v6",
        "MSRPC": "Microsoft Remote Procedure Call",
        "NTP": "Network Time Protocol",
        "TDS": "Tabular Data Stream (SQL Server)",
        "CIFS": "Common Internet File System",
        "006T:dpn": "Custom Protocol",
        "tc107143": "Custom TCP Protocol",
        "tcp8090": "HTTP Alternate on TCP 8090",
        "tcp88": "Kerberos on TCP 88",
        "tcpv143": "IMAP variant",
    }
    
    return port_services, protocol_services

def parse_protocol_port(protocol_str):
    """
    Parse protocol string to extract protocol type and port number.
    Examples: 'udp:26198', 'tcp:443', 'SSL:443', 'HTTP', 'NTP'
    """
    if not protocol_str or pd.isna(protocol_str):
        return None, None, "Unknown"
    
    protocol_str = str(protocol_str).strip()
    
    # Handle port-specific protocols
    if ':' in protocol_str:
        proto_part, port_part = protocol_str.split(':', 1)
        try:
            port = int(port_part)
            if proto_part.upper() in ['UDP', 'TCP', 'SSL']:
                return proto_part.upper(), port, f"{proto_part.upper()}:{port}"
            else:
                return 'TCP', port, f"{proto_part}:{port}"
        except ValueError:
            return None, None, protocol_str
    
    # Handle standalone protocols
    return None, None, protocol_str

def get_service_name(protocol, port, protocol_str, port_services, protocol_services):
    """
    Get service name based on protocol and port information.
    """
    # First check protocol services for exact matches
    if protocol_str in protocol_services:
        return protocol_services[protocol_str]
    
    # Then check port services
    if port and port in port_services:
        return port_services[port]
    
    # Special handling for SSL protocols
    if protocol_str.startswith('SSL:'):
        return f"SSL/TLS Service on port {port if port else 'unknown'}"
    
    # Default based on protocol type
    if protocol == 'UDP':
        return f"UDP Service on port {port}"
    elif protocol == 'TCP':
        return f"TCP Service on port {port}"
    
    return f"Unknown Service ({protocol_str})"

def extract_ip_from_peer(peer_str):
    """
    Extract IP address from peer string.
    Example: '10.165.109.169(upvulnapp11a.unix.rgbk.com)' -> '10.165.109.169'
    """
    if not peer_str or pd.isna(peer_str):
        return None
    
    # Use regex to find IP address pattern
    ip_pattern = r'\b(?:[0-9]{1,3}\.){3}[0-9]{1,3}\b'
    match = re.search(ip_pattern, str(peer_str))
    return match.group() if match else str(peer_str)

def transform_activnet_data(input_file, output_file):
    """
    Transform ACTIVnet data to match synthetic_flows format and add service mapping.
    """
    print("Reading ACTIVnet data...")
    
    # Read the original data
    df_original = pd.read_excel(input_file, sheet_name='app_code_acdm')
    
    print(f"Original data shape: {df_original.shape}")
    print(f"Columns: {list(df_original.columns)}")
    
    # Create port and protocol mappings
    port_services, protocol_services = create_port_service_mapping()
    
    # Create transformed data
    transformed_data = []
    
    for index, row in df_original.iterrows():
        # Parse protocol and port
        protocol, port, protocol_str = parse_protocol_port(row['Protocol'])
        
        # Get service name
        service_name = get_service_name(protocol, port, protocol_str, port_services, protocol_services)
        
        # Extract destination IP
        dst_ip = extract_ip_from_peer(row['Peer'])
        
        # Create timestamp (using current time as example)
        timestamp = int(datetime.now().timestamp())
        
        # Create transformed row matching synthetic_flows format
        transformed_row = {
            'src': row['IP'],
            'dst': dst_ip,
            'port': port if port else '',
            'tier': 'Service',  # Default value, could be customized
            'archetype': 'Network Service',  # Default value, could be customized
            'application': row['Application Name'] if pd.notna(row['Application Name']) else 'Unknown',
            'protocol': protocol if protocol else protocol_str,
            'timestamp': timestamp,
            'info': f"{service_name}",
            'behavior': 'Network Communication',  # Default value
            'application_original': row['Application Name'] if pd.notna(row['Application Name']) else 'Unknown',
            'review_required': False,  # Default value
            'is_known_app': 1 if pd.notna(row['Application Name']) else 0,
            'service_definition': service_name,  # Additional column for service definition
            'bytes_in': row['Bytes In'] if pd.notna(row['Bytes In']) else 0,
            'bytes_out': row['Bytes Out'] if pd.notna(row['Bytes Out']) else 0,
            'original_protocol': row['Protocol'],  # Keep original protocol for reference
            'peer_info': row['Peer'],  # Keep original peer info
            'device_name': row['Name'].strip() if pd.notna(row['Name']) else ''
        }
        
        transformed_data.append(transformed_row)
    
    # Create DataFrame
    df_transformed = pd.DataFrame(transformed_data)
    
    print(f"Transformed data shape: {df_transformed.shape}")
    
    # Load the original workbook
    print("Loading original workbook...")
    workbook = load_workbook(input_file)
    
    # Create new worksheet
    new_sheet_name = 'transformed_flows'
    if new_sheet_name in workbook.sheetnames:
        del workbook[new_sheet_name]
    
    workbook.create_sheet(new_sheet_name)
    worksheet = workbook[new_sheet_name]
    
    # Write headers
    headers = list(df_transformed.columns)
    for col_idx, header in enumerate(headers, 1):
        worksheet.cell(row=1, column=col_idx, value=header)
    
    # Write data
    for row_idx, (_, row) in enumerate(df_transformed.iterrows(), 2):
        for col_idx, value in enumerate(row, 1):
            worksheet.cell(row=row_idx, column=col_idx, value=value)
    
    # Save the workbook
    print(f"Saving transformed data to {output_file}")
    workbook.save(output_file)
    
    # Also save as separate CSV for easy viewing
    csv_output = output_file.replace('.xlsx', '_transformed.csv')
    df_transformed.to_csv(csv_output, index=False)
    
    print(f"Data transformation complete!")
    print(f"New sheet '{new_sheet_name}' added to {output_file}")
    print(f"CSV export saved as {csv_output}")
    
    # Display summary statistics
    print("\n=== TRANSFORMATION SUMMARY ===")
    print(f"Total records processed: {len(df_transformed)}")
    print(f"Unique source IPs: {df_transformed['src'].nunique()}")
    print(f"Unique destination IPs: {df_transformed['dst'].nunique()}")
    print(f"Unique ports: {df_transformed['port'].nunique()}")
    print(f"Unique protocols: {df_transformed['protocol'].nunique()}")
    print(f"Unique applications: {df_transformed['application'].nunique()}")
    
    print("\n=== SERVICE DEFINITIONS FOUND ===")
    service_counts = df_transformed['service_definition'].value_counts()
    for service, count in service_counts.items():
        print(f"{service}: {count} records")
    
    print("\n=== PROTOCOL DISTRIBUTION ===")
    protocol_counts = df_transformed['protocol'].value_counts()
    for protocol, count in protocol_counts.items():
        print(f"{protocol}: {count} records")
    
    return df_transformed

def main():
    """
    Main function to execute the transformation.
    """
    input_file = 'ACTIVnet_Sample.xlsx'
    output_file = 'ACTIVnet_Sample_with_transformed_data.xlsx'
    
    print("ACTIVnet Data Transformer")
    print("=" * 50)
    
    try:
        # Perform the transformation
        df_result = transform_activnet_data(input_file, output_file)
        
        print(f"\n✅ Transformation completed successfully!")
        print(f"📁 Output file: {output_file}")
        print(f"📊 Records processed: {len(df_result)}")
        
        # Display first few rows for verification
        print("\n=== SAMPLE TRANSFORMED DATA ===")
        print(df_result[['src', 'dst', 'port', 'protocol', 'application', 'service_definition']].head(10))
        
    except FileNotFoundError:
        print(f"❌ Error: Input file '{input_file}' not found.")
        print("Please ensure the ACTIVnet_Sample.xlsx file is in the current directory.")
    
    except Exception as e:
        print(f"❌ Error during transformation: {str(e)}")
        print("Please check the input file format and try again.")

if __name__ == "__main__":
    main()