#!/usr/bin/env python3
"""
Dynamic ACTIVnet Data Transformer with Auto Port Research
Automatically researches unknown ports using multiple online databases and APIs.
Includes intelligent caching and fallback mechanisms.
"""

import pandas as pd
import re
import json
import time
import requests
from datetime import datetime
import openpyxl
from openpyxl import load_workbook
from pathlib import Path
import urllib.parse
from bs4 import BeautifulSoup
import pickle
import hashlib

class PortResearcher:
    """
    Dynamic port research class that automatically looks up unknown ports
    from multiple online sources with intelligent caching.
    """
    
    def __init__(self, cache_file='port_cache.json'):
        self.cache_file = cache_file
        self.cache = self.load_cache()
        self.session = requests.Session()
        self.session.headers.update({
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
        })
        
        # Basic well-known ports for quick lookup (minimal set)
        self.well_known_ports = {
            7: "Echo Protocol",
            21: "FTP (File Transfer Protocol)",
            22: "SSH (Secure Shell)", 
            23: "Telnet",
            25: "SMTP (Simple Mail Transfer Protocol)",
            53: "DNS (Domain Name System)",
            80: "HTTP (HyperText Transfer Protocol)",
            123: "NTP (Network Time Protocol)",
            443: "HTTPS (HTTP over SSL/TLS)",
            445: "SMB (Server Message Block)",
        }
        
        # Protocol service mappings
        self.protocol_services = {
            "DNS": "Domain Name System",
            "HTTP": "HyperText Transfer Protocol",
            "HTTP2-SSL": "HTTP/2 over SSL/TLS",
            "HTTPS": "HTTP over SSL/TLS",
            "HTTps": "HTTP over SSL/TLS (variant)",
            "ICMPv6": "Internet Control Message Protocol v6",
            "MSRPC": "Microsoft Remote Procedure Call",
            "NTP": "Network Time Protocol",
            "TDS": "Tabular Data Stream (SQL Server)",
            "CIFS": "Common Internet File System",
        }

    def load_cache(self):
        """Load cached port information from file."""
        try:
            if Path(self.cache_file).exists():
                with open(self.cache_file, 'r') as f:
                    return json.load(f)
        except Exception as e:
            print(f"Warning: Could not load cache file: {e}")
        return {}

    def save_cache(self):
        """Save port information to cache file."""
        try:
            with open(self.cache_file, 'w') as f:
                json.dump(self.cache, f, indent=2)
        except Exception as e:
            print(f"Warning: Could not save cache file: {e}")

    def get_cache_key(self, port, protocol):
        """Generate cache key for port/protocol combination."""
        return f"{protocol}:{port}" if protocol and port else str(port or protocol)

    def research_port_iana(self, port, protocol):
        """
        Research port using IANA service registry.
        """
        try:
            # IANA CSV endpoint for port registry
            url = "https://www.iana.org/assignments/service-names-port-numbers/service-names-port-numbers.csv"
            response = self.session.get(url, timeout=10)
            
            if response.status_code == 200:
                lines = response.text.split('\n')
                for line in lines[1:]:  # Skip header
                    parts = [p.strip('"') for p in line.split(',')]
                    if len(parts) >= 4:
                        service_name = parts[0]
                        port_num = parts[1]
                        transport = parts[2]
                        description = parts[3]
                        
                        if (port_num == str(port) and 
                            (not protocol or transport.upper() == protocol.upper())):
                            return f"{service_name} - {description}".strip(' -')
                            
        except Exception as e:
            print(f"IANA lookup failed for {protocol}:{port} - {e}")
        return None

    def research_port_speedguide(self, port, protocol):
        """
        Research port using SpeedGuide port database.
        """
        try:
            url = f"https://www.speedguide.net/port.php?port={port}"
            response = self.session.get(url, timeout=10)
            
            if response.status_code == 200:
                soup = BeautifulSoup(response.content, 'html.parser')
                
                # Look for service information in the table
                tables = soup.find_all('table')
                for table in tables:
                    rows = table.find_all('tr')
                    for row in rows:
                        cells = row.find_all('td')
                        if len(cells) >= 4:
                            port_cell = cells[0].get_text().strip()
                            protocol_cell = cells[1].get_text().strip()
                            service_cell = cells[2].get_text().strip()
                            details_cell = cells[3].get_text().strip()
                            
                            if (port_cell == str(port) and 
                                (not protocol or protocol.lower() in protocol_cell.lower())):
                                service_info = f"{service_cell}"
                                if details_cell and details_cell != service_cell:
                                    service_info += f" - {details_cell}"
                                return service_info[:200]  # Limit length
                                
        except Exception as e:
            print(f"SpeedGuide lookup failed for {protocol}:{port} - {e}")
        return None

    def research_port_whatportis(self, port, protocol):
        """
        Research port using WhatPortIs database.
        """
        try:
            url = f"https://whatportis.com/port/{port}"
            response = self.session.get(url, timeout=10)
            
            if response.status_code == 200:
                soup = BeautifulSoup(response.content, 'html.parser')
                
                # Look for service descriptions
                service_divs = soup.find_all('div', class_='service-info')
                if service_divs:
                    service_text = service_divs[0].get_text(strip=True)
                    return service_text[:200]  # Limit length
                    
                # Alternative: look for any text containing service info
                text_content = soup.get_text()
                if f"port {port}" in text_content.lower():
                    lines = text_content.split('\n')
                    for line in lines:
                        if f"port {port}" in line.lower() and len(line) < 200:
                            return line.strip()
                            
        except Exception as e:
            print(f"WhatPortIs lookup failed for {protocol}:{port} - {e}")
        return None

    def research_port_sans(self, port, protocol):
        """
        Research port using SANS Internet Storm Center.
        """
        try:
            url = f"https://isc.sans.edu/data/port/{port}"
            response = self.session.get(url, timeout=10)
            
            if response.status_code == 200:
                soup = BeautifulSoup(response.content, 'html.parser')
                
                # Look for port information
                content = soup.get_text()
                if "Description:" in content:
                    lines = content.split('\n')
                    for i, line in enumerate(lines):
                        if "Description:" in line and i + 1 < len(lines):
                            desc = lines[i + 1].strip()
                            if desc and len(desc) < 200:
                                return desc
                                
        except Exception as e:
            print(f"SANS lookup failed for {protocol}:{port} - {e}")
        return None

    def research_port_online(self, port, protocol):
        """
        Research port information using multiple online sources.
        """
        cache_key = self.get_cache_key(port, protocol)
        
        # Check cache first
        if cache_key in self.cache:
            print(f"  📋 Found {protocol}:{port} in cache")
            return self.cache[cache_key]

        print(f"  🔍 Researching {protocol}:{port} online...")
        
        # Try multiple sources in order of reliability
        sources = [
            ("IANA", self.research_port_iana),
            ("SpeedGuide", self.research_port_speedguide),
            ("WhatPortIs", self.research_port_whatportis),
            ("SANS", self.research_port_sans),
        ]
        
        for source_name, research_func in sources:
            try:
                result = research_func(port, protocol)
                if result and result.strip():
                    service_info = result.strip()
                    print(f"    ✅ Found via {source_name}: {service_info}")
                    
                    # Cache the result
                    self.cache[cache_key] = service_info
                    self.save_cache()
                    
                    return service_info
                    
                # Rate limiting - be respectful
                time.sleep(0.5)
                
            except Exception as e:
                print(f"    ❌ {source_name} failed: {e}")
                continue

        # If no online source found anything, create a reasonable default
        default_service = f"Unknown {protocol.upper()} Service on port {port}" if protocol and port else f"Unknown Service ({protocol or port})"
        
        # Cache the default to avoid repeated lookups
        self.cache[cache_key] = default_service
        self.save_cache()
        
        print(f"    ⚠️  No information found, using default: {default_service}")
        return default_service

    def get_service_name(self, protocol, port, protocol_str):
        """
        Get service name with automatic research for unknown ports.
        """
        # Check protocol services first
        if protocol_str in self.protocol_services:
            return self.protocol_services[protocol_str]
        
        # Check well-known ports for quick response
        if port and port in self.well_known_ports:
            return self.well_known_ports[port]
        
        # Special handling for SSL protocols
        if protocol_str and protocol_str.startswith('SSL:'):
            try:
                ssl_port = int(protocol_str.split(':')[1])
                return f"SSL/TLS Service on port {ssl_port}"
            except:
                return f"SSL/TLS Service ({protocol_str})"
        
        # For unknown ports, research online
        if port and port not in self.well_known_ports:
            return self.research_port_online(port, protocol)
        
        # Default fallback
        return f"Unknown Service ({protocol_str})"

class DataTransformer:
    """
    Main data transformation class.
    """
    
    def __init__(self):
        self.port_researcher = PortResearcher()

    def parse_protocol_port(self, protocol_str):
        """
        Parse protocol string to extract protocol type and port number.
        """
        if not protocol_str or pd.isna(protocol_str):
            return None, None, "Unknown"
        
        protocol_str = str(protocol_str).strip()
        
        # Handle port-specific protocols
        if ':' in protocol_str:
            proto_part, port_part = protocol_str.split(':', 1)
            try:
                port = int(port_part)
                if proto_part.upper() in ['UDP', 'TCP', 'SSL']:
                    return proto_part.upper(), port, f"{proto_part.upper()}:{port}"
                else:
                    return 'TCP', port, f"{proto_part}:{port}"
            except ValueError:
                return None, None, protocol_str
        
        return None, None, protocol_str

    def extract_ip_from_peer(self, peer_str):
        """
        Extract IP address from peer string.
        """
        if not peer_str or pd.isna(peer_str):
            return None
        
        ip_pattern = r'\b(?:[0-9]{1,3}\.){3}[0-9]{1,3}\b'
        match = re.search(ip_pattern, str(peer_str))
        return match.group() if match else str(peer_str)

    def transform_activnet_data(self, input_file, output_file):
        """
        Transform ACTIVnet data with automatic port research.
        """
        print("🚀 Starting ACTIVnet Data Transformation with Auto Port Research")
        print("=" * 70)
        
        # Read the original data
        print("📖 Reading ACTIVnet data...")
        df_original = pd.read_excel(input_file, sheet_name='app_code_acdm')
        
        print(f"📊 Original data shape: {df_original.shape}")
        print(f"📋 Columns: {list(df_original.columns)}")
        
        # Analyze unique protocols first
        unique_protocols = df_original['Protocol'].unique()
        print(f"\n🔍 Found {len(unique_protocols)} unique protocols")
        print("🔬 Pre-analyzing protocols for research...")
        
        # Pre-research unique ports to minimize API calls
        unique_ports = set()
        for protocol_str in unique_protocols:
            if pd.notna(protocol_str):
                protocol, port, _ = self.parse_protocol_port(protocol_str)
                if port:
                    unique_ports.add((protocol, port))
        
        print(f"🎯 Will research {len(unique_ports)} unique port combinations")
        
        # Transform data
        print("\n🔄 Transforming data...")
        transformed_data = []
        
        for index, row in df_original.iterrows():
            if index % 20 == 0:
                print(f"  📍 Processing row {index + 1}/{len(df_original)}")
            
            # Parse protocol and port
            protocol, port, protocol_str = self.parse_protocol_port(row['Protocol']) 
            
            # Get service name (with automatic research)
            service_name = self.port_researcher.get_service_name(protocol, port, protocol_str)
            
            # Extract destination IP
            dst_ip = self.extract_ip_from_peer(row['Peer'])
            
            # Create timestamp
            timestamp = int(datetime.now().timestamp())
            
            # Create transformed row
            transformed_row = {
                'src': row['IP'],
                'dst': dst_ip,
                'port': port if port else '',
                'tier': 'Service',
                'archetype': 'Network Service',
                'application': row['Application Name'] if pd.notna(row['Application Name']) else 'Unknown',
                'protocol': protocol if protocol else protocol_str,
                'timestamp': timestamp,
                'info': f"{service_name}",
                'behavior': 'Network Communication',
                'application_original': row['Application Name'] if pd.notna(row['Application Name']) else 'Unknown',
                'review_required': False,
                'is_known_app': 1 if pd.notna(row['Application Name']) else 0,
                'service_definition': service_name,
                'bytes_in': row['Bytes In'] if pd.notna(row['Bytes In']) else 0,
                'bytes_out': row['Bytes Out'] if pd.notna(row['Bytes Out']) else 0,
                'original_protocol': row['Protocol'],
                'peer_info': row['Peer'],
                'device_name': row['Name'].strip() if pd.notna(row['Name']) else ''
            }
            
            transformed_data.append(transformed_row)
        
        # Create DataFrame
        df_transformed = pd.DataFrame(transformed_data)
        
        print(f"\n✅ Data transformation completed!")
        print(f"📊 Transformed data shape: {df_transformed.shape}")
        
        # Save results
        print("\n💾 Saving results...")
        self.save_results(input_file, output_file, df_transformed)
        
        # Print summary
        self.print_summary(df_transformed)
        
        return df_transformed

    def save_results(self, input_file, output_file, df_transformed):
        """
        Save the transformed results to Excel and CSV.
        """
        # Load the original workbook
        workbook = load_workbook(input_file)
        
        # Create new worksheet
        new_sheet_name = 'transformed_flows_auto_research'
        if new_sheet_name in workbook.sheetnames:
            del workbook[new_sheet_name]
        
        workbook.create_sheet(new_sheet_name)
        worksheet = workbook[new_sheet_name]
        
        # Write headers
        headers = list(df_transformed.columns)
        for col_idx, header in enumerate(headers, 1):
            worksheet.cell(row=1, column=col_idx, value=header)
        
        # Write data
        for row_idx, (_, row) in enumerate(df_transformed.iterrows(), 2):
            for col_idx, value in enumerate(row, 1):
                worksheet.cell(row=row_idx, column=col_idx, value=value)
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Save the workbook
        workbook.save(output_file)
        
        # Also save as CSV
        csv_output = output_file.replace('.xlsx', '_auto_research.csv')
        df_transformed.to_csv(csv_output, index=False)
        
        print(f"📁 Excel saved: {output_file}")
        print(f"📁 CSV saved: {csv_output}")

    def print_summary(self, df_transformed):
        """
        Print transformation summary and statistics.
        """
        print("\n" + "=" * 50)
        print("📈 TRANSFORMATION SUMMARY")
        print("=" * 50)
        print(f"Total records processed: {len(df_transformed)}")
        print(f"Unique source IPs: {df_transformed['src'].nunique()}")
        print(f"Unique destination IPs: {df_transformed['dst'].nunique()}")
        print(f"Unique ports: {df_transformed['port'].nunique()}")
        print(f"Unique protocols: {df_transformed['protocol'].nunique()}")
        print(f"Unique applications: {df_transformed['application'].nunique()}")
        
        print("\n🔍 SERVICE DEFINITIONS DISCOVERED")
        print("-" * 30)
        service_counts = df_transformed['service_definition'].value_counts()
        for service, count in service_counts.head(15).items():
            print(f"{service}: {count} records")
        
        print(f"\n📡 PROTOCOL DISTRIBUTION")
        print("-" * 20)
        protocol_counts = df_transformed['protocol'].value_counts()
        for protocol, count in protocol_counts.items():
            print(f"{protocol}: {count} records")
        
        # Show cache statistics
        cache_size = len(self.port_researcher.cache)
        print(f"\n💾 CACHE STATISTICS")
        print(f"Cached port definitions: {cache_size}")
        print(f"Cache file: {self.port_researcher.cache_file}")

def main():
    """
    Main execution function.
    """
    input_file = 'ACTIVnet_Sample.xlsx'
    output_file = 'ACTIVnet_Sample_with_auto_research.xlsx'
    
    print("🎯 Dynamic ACTIVnet Data Transformer")
    print("🤖 With Automatic Port Research & Caching")
    print("=" * 60)
    
    try:
        # Check for required libraries
        try:
            import requests
            from bs4 import BeautifulSoup
        except ImportError:
            print("❌ Missing required libraries!")
            print("Please install: pip install requests beautifulsoup4 pandas openpyxl")
            return
        
        # Create transformer and run
        transformer = DataTransformer()
        df_result = transformer.transform_activnet_data(input_file, output_file)
        
        print(f"\n🎉 TRANSFORMATION COMPLETED SUCCESSFULLY!")
        print(f"📁 Output file: {output_file}")
        print(f"📊 Records processed: {len(df_result)}")
        print(f"🔍 Auto-researched unknown ports using multiple online databases")
        print(f"💾 Results cached for future runs")
        
        # Display sample results
        print(f"\n📋 SAMPLE TRANSFORMED DATA")
        print("-" * 40)
        sample_cols = ['src', 'dst', 'port', 'protocol', 'application', 'service_definition']
        print(df_result[sample_cols].head(10).to_string(index=False))
        
    except FileNotFoundError:
        print(f"❌ Error: Input file '{input_file}' not found.")
        print("Please ensure the ACTIVnet_Sample.xlsx file is in the current directory.")
    
    except Exception as e:
        print(f"❌ Error during transformation: {str(e)}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()